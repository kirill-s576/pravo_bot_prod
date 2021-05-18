# Django Environment activation stage
import os
import django
from dotenv import load_dotenv
from pathlib import Path
env_path = Path('/Users/kirill/own_projects/pravo_bot/environments.env')
load_dotenv(dotenv_path=env_path)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()
#####

from quiz.models import *
import openpyxl


## Button loader
# wb = openpyxl.load_workbook("/Users/kirill/own_projects/pravo_bot/utils/eng_btn.xlsx")
# ws = wb.active
#
# max_row = ws.max_row
# lang = Language.objects.get(label="EN")
#
#
# for row in range(2, max_row+1):
#     i = int(ws.cell(row, 1).value)
#     value = ws.cell(row, 2).value
#     button = QButton.objects.get(id=i)
#     QButtonTranslation.objects.create(
#         language=lang,
#         button=button,
#         text=value
#     )

# Msg loader
wb = openpyxl.load_workbook("/Users/kirill/own_projects/pravo_bot/utils/arabic_msg.xlsx")
ws = wb.active

max_row = ws.max_row
lang = Language.objects.get(label="AR")

MessageTranslation.objects.filter(language=lang).delete()

for row in range(2, max_row + 1):
    i = int(ws.cell(row, 1).value)
    value = ws.cell(row, 2).value
    message = Message.objects.get(id=i)
    MessageTranslation.objects.create(
        language=lang,
        message=message,
        text=value
    )
